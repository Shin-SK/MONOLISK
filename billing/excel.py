# billing/excel.py
"""Excel出力ロジック（個別伝票 / 日次ZIP / 売上日報）"""
import re
import zipfile
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.utils import timezone


# ── 共通スタイル ──
HEADER_FONT = Font(bold=True, size=10)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _fmt_dt(dt):
    """datetime → 文字列。Noneなら空文字"""
    if dt is None:
        return ""
    local = timezone.localtime(dt)
    return local.strftime("%Y-%m-%d %H:%M")


def _fmt_time(dt):
    if dt is None:
        return ""
    local = timezone.localtime(dt)
    return local.strftime("%H:%M")


def _set_header_row(ws, row, headers):
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")


def _set_data_row(ws, row, values):
    for col_idx, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=v)
        cell.border = THIN_BORDER


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 30)


def _get_cast_names(bill_item):
    """BillItemの担当者名を取得。M2M優先、なければFK、なければ'担当なし'"""
    # M2M (served_by_casts)
    casts_m2m = list(bill_item.served_by_casts.all())
    if casts_m2m:
        return ", ".join(c.stage_name for c in casts_m2m)
    # FK (served_by_cast)
    if bill_item.served_by_cast:
        return bill_item.served_by_cast.stage_name
    return "担当なし"


# ============================================================
#  個別伝票Excel
# ============================================================
def generate_bill_excel(bill):
    """1伝票のExcelをBytesIOで返す"""
    wb = Workbook()
    ws = wb.active
    ws.title = "伝票"

    # ─── ヘッダ情報 ───
    row = 1
    ws.cell(row=row, column=1, value="伝票").font = TITLE_FONT
    row += 2

    # 基本情報
    info_labels = ["伝票番号", "営業日", "開始", "終了", "テーブル", "顧客", "人数", "メモ"]
    tables_str = ", ".join(t.code or str(t.id) for t in bill.tables.all()) or (bill.table.code if bill.table else "")
    customers = list(bill.customers.all())
    customers_str = ", ".join(c.alias or c.full_name for c in customers) if customers else ""
    display = bill.display_name or ""

    info_values = [
        f"#{bill.id}" + (f" {display}" if display else ""),
        _fmt_dt(bill.opened_at).split(" ")[0] if bill.opened_at else "",
        _fmt_dt(bill.opened_at),
        _fmt_dt(bill.closed_at),
        tables_str,
        customers_str,
        bill.pax,
        bill.memo or "",
    ]

    for i, (label, value) in enumerate(zip(info_labels, info_values)):
        r = row + i
        cell_l = ws.cell(row=r, column=1, value=label)
        cell_l.font = HEADER_FONT
        cell_l.fill = HEADER_FILL
        cell_l.border = THIN_BORDER
        cell_v = ws.cell(row=r, column=2, value=value)
        cell_v.border = THIN_BORDER
    row += len(info_labels) + 1

    # ─── 注文明細 ───
    ws.cell(row=row, column=1, value="注文明細").font = Font(bold=True, size=12)
    row += 1

    order_headers = ["注文", "注文日時", "担当", "個数", "計"]
    _set_header_row(ws, row, order_headers)
    row += 1

    items = bill.items.select_related("served_by_cast", "item_master").prefetch_related("served_by_casts").all()
    for item in items:
        cast_names = _get_cast_names(item)
        values = [
            item.name,
            _fmt_dt(item.ordered_at),
            cast_names,
            item.qty,
            item.subtotal,
        ]
        _set_data_row(ws, row, values)
        row += 1

    row += 1

    # ─── 立替明細 ───
    ws.cell(row=row, column=1, value="立替明細").font = Font(bold=True, size=12)
    row += 1

    sub_headers = ["立替", "日時", "対象キャスト", "個数", "客引き額", "給与控除額"]
    _set_header_row(ws, row, sub_headers)
    row += 1

    subs = bill.substitute_items.select_related("cast").all()
    for s in subs:
        values = [
            s.name,
            _fmt_dt(s.ordered_at),
            s.cast.stage_name if s.cast else "",
            s.qty,
            (s.price or 0) * (s.qty or 1),
            s.substitute_amount,
        ]
        _set_data_row(ws, row, values)
        row += 1

    row += 1

    # ─── 会計 ───
    ws.cell(row=row, column=1, value="会計").font = Font(bold=True, size=12)
    row += 1

    # 立替による控除合計
    substitute_total = sum(
        (si.price or 0) * (si.qty or 1)
        for si in bill.substitute_items.all()
    )

    accounting = [
        ("注文小計", bill.subtotal),
        ("立替による控除", substitute_total),
        ("サービス料", bill.service_charge),
        ("税", bill.tax),
        ("総額", bill.grand_total),
    ]

    for label, value in accounting:
        cell_l = ws.cell(row=row, column=1, value=label)
        cell_l.font = HEADER_FONT
        cell_l.fill = HEADER_FILL
        cell_l.border = THIN_BORDER
        cell_v = ws.cell(row=row, column=2, value=value)
        cell_v.border = THIN_BORDER
        cell_v.number_format = "#,##0"
        row += 1

    row += 1

    # ─── 稼働キャスト ───
    ws.cell(row=row, column=1, value="稼働キャスト").font = Font(bold=True, size=12)
    row += 1

    stay_headers = ["キャスト", "種別", "入店", "退店"]
    _set_header_row(ws, row, stay_headers)
    row += 1

    STAY_TYPE_LABELS = {"free": "フリー", "in": "場内指名", "nom": "本指名", "dohan": "同伴"}
    stays = bill.stays.select_related("cast").all()
    for stay in stays:
        values = [
            stay.cast.stage_name if stay.cast else "",
            STAY_TYPE_LABELS.get(stay.stay_type, stay.stay_type),
            _fmt_time(stay.entered_at),
            _fmt_time(stay.left_at),
        ]
        _set_data_row(ws, row, values)
        row += 1

    _auto_width(ws)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _safe_filename(s):
    """ファイル名に使えない文字を除去"""
    return re.sub(r'[\\/:*?"<>|\s]+', '_', s).strip('_') or 'unknown'


# ============================================================
#  1日分まとめZIP
# ============================================================
def generate_daily_zip(bills):
    """伝票リストからZIP(BytesIO)を生成。billsが空ならNoneを返す"""
    if not bills:
        return None

    zip_buf = BytesIO()
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for bill in bills:
            # テーブルコード
            tables = list(bill.tables.all())
            if tables:
                table_code = '_'.join(_safe_filename(t.code or str(t.id)) for t in tables)
            elif bill.table:
                table_code = _safe_filename(bill.table.code or str(bill.table.id))
            else:
                table_code = 'notable'

            # 開始時刻
            time_str = timezone.localtime(bill.opened_at).strftime('%H%M') if bill.opened_at else '0000'

            fname = f'bill_{bill.id}_{table_code}_{time_str}.xlsx'
            excel_buf = generate_bill_excel(bill)
            zf.writestr(fname, excel_buf.getvalue())

    zip_buf.seek(0)
    return zip_buf


# ============================================================
#  売上日報Excel
# ============================================================
REPORT_COLUMNS = [
    "No", "テーブル", "顧客", "人数", "開始", "終了", "担当",
    "伝票番号", "注文小計", "立替控除", "サービス料", "税", "総額",
    "現金", "カード", "カード種別", "メモ",
]

# 合計対象の列インデックス（0始まり）
_SUM_COLS = {3, 8, 9, 10, 11, 12, 13, 14}  # 人数,注文小計,立替控除,サービス料,税,総額,現金,カード

SUM_FONT = Font(bold=True, size=10)
SUM_BORDER = Border(top=Side(style="medium"), bottom=Side(style="medium"),
                    left=Side(style="thin"), right=Side(style="thin"))


def _bill_cast_str(bill):
    """日報用の担当キャスト名。main_cast優先、なければstaysから重複排除"""
    if bill.main_cast:
        return bill.main_cast.stage_name
    names = []
    seen = set()
    for stay in bill.stays.all():
        if stay.cast_id and stay.cast_id not in seen:
            seen.add(stay.cast_id)
            names.append(stay.cast.stage_name)
    return ", ".join(names)


def _bill_table_str(bill):
    tables = list(bill.tables.all())
    if tables:
        return ", ".join(t.code or str(t.id) for t in tables)
    if bill.table:
        return bill.table.code or str(bill.table.id)
    return ""


def generate_daily_report(bills, target_date):
    """売上日報ExcelをBytesIOで返す"""
    wb = Workbook()
    ws = wb.active
    ws.title = "売上日報"

    # ─── タイトル ───
    row = 1
    title_cell = ws.cell(row=row, column=1, value=f"売上日報 {target_date}")
    title_cell.font = TITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(REPORT_COLUMNS))
    row += 2

    # ─── ヘッダ ───
    _set_header_row(ws, row, REPORT_COLUMNS)
    row += 1

    # ─── データ行 ───
    sums = {i: 0 for i in _SUM_COLS}

    for idx, bill in enumerate(bills, 1):
        customers = list(bill.customers.all())
        customers_str = ", ".join(c.alias or c.full_name for c in customers) if customers else ""

        substitute_deduction = sum(
            (si.price or 0) * (si.qty or 1)
            for si in bill.substitute_items.all()
        )

        values = [
            idx,                                          # No
            _bill_table_str(bill),                        # テーブル
            customers_str,                                # 顧客
            bill.pax,                                     # 人数
            _fmt_time(bill.opened_at),                    # 開始
            _fmt_time(bill.closed_at),                    # 終了
            _bill_cast_str(bill),                             # 担当
            bill.id,                                      # 伝票番号
            bill.subtotal,                                # 注文小計
            substitute_deduction,                         # 立替控除
            bill.service_charge,                          # サービス料
            bill.tax,                                     # 税
            bill.grand_total,                             # 総額
            bill.paid_cash,                               # 現金
            bill.paid_card,                               # カード
            bill.get_card_brand_display() if bill.card_brand else "",  # カード種別
            bill.memo or "",                              # メモ
        ]

        _set_data_row(ws, row, values)
        # 金額列に書式
        for ci in _SUM_COLS:
            ws.cell(row=row, column=ci + 1).number_format = "#,##0"
            sums[ci] += (values[ci] or 0)
        row += 1

    # ─── 合計行 ───
    row += 0  # データ直下
    for col_idx in range(len(REPORT_COLUMNS)):
        cell = ws.cell(row=row, column=col_idx + 1)
        cell.border = SUM_BORDER
        cell.font = SUM_FONT
        if col_idx == 0:
            cell.value = "合計"
        elif col_idx in _SUM_COLS:
            cell.value = sums[col_idx]
            cell.number_format = "#,##0"

    _auto_width(ws)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
